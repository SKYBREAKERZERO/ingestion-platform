from __future__ import annotations

import math
import zipfile
from xml.etree.ElementTree import iterparse
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.worksheet.worksheet import Worksheet

from app.loader.base_loader import BaseLoader
from app.model.block import BlockType, DocumentBlock
from app.model.document import Document
from app.model.page import Page


class XLSXLoaderError(RuntimeError):
    """XLSX 文件加载异常。"""


class XLSXLoader(BaseLoader):
    """
    XLSX 原始内容加载器。

    负责：
        - 加载 XLSX 工作簿
        - 识别实际有值单元格范围，规避 Excel UsedRange / 格式污染
        - 按工作表和行顺序读取单元格
        - 每个非空行生成一个 TABLE Block
        - 保留 Sheet、行、列和单元格位置
        - 生成兼容现有架构的逻辑 Page
        - 写入 Loader 元数据

    不负责：
        - Chapter / Section 建模
        - 表头识别
        - 多表区域识别
        - 行过滤
        - Chunk
        - Token 统计
        - JSON / PostgreSQL 保存

    默认映射：
        Worksheet -> 逻辑 Page
        Non-empty Row -> DocumentBlock(TABLE)
    """

    SUPPORTED_EXTENSIONS = {
        ".xlsx",
    }

    def __init__(
        self,
        *,
        read_only: bool = True,
        data_only: bool = False,
        keep_links: bool = False,
        include_hidden_sheets: bool = True,
        include_very_hidden_sheets: bool = True,
        include_empty_cells_between_values: bool = True,
        preserve_formulas_in_metadata: bool = False,
        row_separator: str = " | ",
        maximum_rows_per_sheet: int | None = None,
        maximum_columns_per_sheet: int | None = None,
        trim_inflated_used_range: bool = True,
        used_range_inflation_ratio: float = 4.0,
        used_range_inflation_min_extra_rows: int = 1000,
        print_summary: bool = False,
    ) -> None:

        if not row_separator:
            raise ValueError(
                "row_separator cannot be empty."
            )

        if (
            maximum_rows_per_sheet is not None
            and maximum_rows_per_sheet <= 0
        ):
            raise ValueError(
                "maximum_rows_per_sheet must be greater than 0."
            )

        if (
            maximum_columns_per_sheet is not None
            and maximum_columns_per_sheet <= 0
        ):
            raise ValueError(
                "maximum_columns_per_sheet must be greater than 0."
            )

        if used_range_inflation_ratio < 1.0:
            raise ValueError(
                "used_range_inflation_ratio must be at least 1.0."
            )

        if used_range_inflation_min_extra_rows < 0:
            raise ValueError(
                "used_range_inflation_min_extra_rows cannot be negative."
            )

        if preserve_formulas_in_metadata and read_only:
            raise ValueError(
                "preserve_formulas_in_metadata=True requires "
                "read_only=False."
            )

        if preserve_formulas_in_metadata and data_only:
            raise ValueError(
                "preserve_formulas_in_metadata=True requires "
                "data_only=False."
            )

        self.read_only = read_only
        self.data_only = data_only
        self.keep_links = keep_links

        self.include_hidden_sheets = (
            include_hidden_sheets
        )

        self.include_very_hidden_sheets = (
            include_very_hidden_sheets
        )

        self.include_empty_cells_between_values = (
            include_empty_cells_between_values
        )

        self.preserve_formulas_in_metadata = (
            preserve_formulas_in_metadata
        )

        self.row_separator = row_separator

        self.maximum_rows_per_sheet = (
            maximum_rows_per_sheet
        )

        self.maximum_columns_per_sheet = (
            maximum_columns_per_sheet
        )

        self.trim_inflated_used_range = bool(
            trim_inflated_used_range
        )

        self.used_range_inflation_ratio = float(
            used_range_inflation_ratio
        )

        self.used_range_inflation_min_extra_rows = int(
            used_range_inflation_min_extra_rows
        )

        self.print_summary = bool(
            print_summary
        )

    def load(
        self,
        file_path: str | Path,
    ) -> Document:

        path = self._validate_path(
            file_path
        )

        workbook = None

        try:
            workbook = load_workbook(
                filename=str(path),
                read_only=self.read_only,
                data_only=self.data_only,
                keep_links=self.keep_links,
            )

            blocks: list[DocumentBlock] = []
            pages: list[Page] = []

            order = 0

            total_sheet_count = len(
                workbook.worksheets
            )

            processed_sheet_count = 0
            skipped_sheet_count = 0
            empty_sheet_count = 0

            total_row_count = 0
            total_non_empty_row_count = 0
            total_cell_count = 0
            total_non_empty_cell_count = 0

            used_range_trimmed_sheet_count = 0
            used_range_total_raw_max_row = 0
            used_range_total_effective_max_row = 0

            sheet_metadata: list[
                dict[str, Any]
            ] = []

            for sheet_index, worksheet in enumerate(
                workbook.worksheets
            ):
                sheet_state = self._get_sheet_state(
                    worksheet
                )

                if not self._should_include_sheet(
                    sheet_state
                ):
                    skipped_sheet_count += 1

                    sheet_metadata.append(
                        {
                            "sheet_index": sheet_index,
                            "sheet_name": worksheet.title,
                            "sheet_state": sheet_state,
                            "status": "SKIPPED",
                            "reason": (
                                "hidden_sheet"
                                if sheet_state == "hidden"
                                else "very_hidden_sheet"
                            ),
                        }
                    )

                    continue

                processed_sheet_count += 1

                sheet_lines: list[str] = []

                sheet_row_count = 0
                sheet_non_empty_row_count = 0
                sheet_cell_count = 0
                sheet_non_empty_cell_count = 0

                first_data_row: int | None = None
                last_data_row: int | None = None
                first_data_column: int | None = None
                last_data_column: int | None = None

                raw_max_row = int(
                    getattr(
                        worksheet,
                        "max_row",
                        0,
                    )
                    or 0
                )

                raw_max_column = int(
                    getattr(
                        worksheet,
                        "max_column",
                        0,
                    )
                    or 0
                )

                (
                    effective_max_row,
                    effective_max_column,
                    used_range_diagnostics,
                ) = self._resolve_effective_used_range(
                    workbook_path=path,
                    worksheet=worksheet,
                    raw_max_row=raw_max_row,
                    raw_max_column=raw_max_column,
                )

                used_range_total_raw_max_row += (
                    raw_max_row
                )

                used_range_total_effective_max_row += (
                    effective_max_row
                )

                if used_range_diagnostics[
                    "trimmed"
                ]:

                    used_range_trimmed_sheet_count += 1

                row_iterator = self._iter_rows(
                    worksheet,
                    max_row=(
                        effective_max_row
                    ),
                    max_column=(
                        effective_max_column
                    ),
                )

                for row_number, row in enumerate(
                    row_iterator,
                    start=1,
                ):
                    if (
                        self.maximum_rows_per_sheet
                        is not None
                        and row_number
                        > self.maximum_rows_per_sheet
                    ):
                        break

                    sheet_row_count += 1
                    total_row_count += 1

                    (
                        values,
                        coordinates,
                        formulas,
                        raw_cell_count,
                        non_empty_cell_count,
                    ) = self._extract_row(
                        row=row,
                        row_number=row_number,
                    )

                    sheet_cell_count += raw_cell_count
                    total_cell_count += raw_cell_count

                    sheet_non_empty_cell_count += (
                        non_empty_cell_count
                    )

                    total_non_empty_cell_count += (
                        non_empty_cell_count
                    )

                    if not values:
                        continue

                    sheet_non_empty_row_count += 1
                    total_non_empty_row_count += 1

                    non_empty_columns = [
                        index
                        for index, value
                        in enumerate(
                            values,
                            start=1,
                        )
                        if value != ""
                    ]

                    if non_empty_columns:
                        row_first_column = min(
                            non_empty_columns
                        )

                        row_last_column = max(
                            non_empty_columns
                        )

                        first_data_column = (
                            row_first_column
                            if first_data_column is None
                            else min(
                                first_data_column,
                                row_first_column,
                            )
                        )

                        last_data_column = (
                            row_last_column
                            if last_data_column is None
                            else max(
                                last_data_column,
                                row_last_column,
                            )
                        )

                    if first_data_row is None:
                        first_data_row = row_number

                    last_data_row = row_number

                    row_text = self.row_separator.join(
                        values
                    )

                    block_metadata: dict[str, Any] = {
                        "source": "xlsx",
                        "sheet_name": worksheet.title,
                        "sheet_index": sheet_index,
                        "sheet_state": sheet_state,
                        "row_number": row_number,
                        "column_count": len(values),
                        "cell_coordinates": coordinates,
                        "column_position_preserved": (
                            self.include_empty_cells_between_values
                        ),
                    }

                    if formulas:
                        block_metadata[
                            "formulas"
                        ] = formulas

                    block = DocumentBlock(
                        id=(
                            f"xlsx-"
                            f"{sheet_index:04d}-"
                            f"row-{row_number:08d}"
                        ),
                        block_type=BlockType.TABLE,
                        text=row_text,
                        order=order,
                        page_number=(
                            processed_sheet_count
                        ),
                        table_index=sheet_index,
                        row_index=row_number - 1,
                        cells=values,
                        source="xlsx",
                        metadata=block_metadata,
                    )

                    blocks.append(block)
                    sheet_lines.append(row_text)

                    order += 1

                if not sheet_lines:
                    empty_sheet_count += 1

                logical_page_number = (
                    len(pages) + 1
                )

                pages.append(
                    Page(
                        page_number=logical_page_number,
                        text="\n".join(
                            sheet_lines
                        ).strip(),
                    )
                )

                sheet_metadata.append(
                    {
                        "sheet_index": sheet_index,
                        "sheet_name": worksheet.title,
                        "sheet_state": sheet_state,
                        "status": (
                            "EMPTY"
                            if not sheet_lines
                            else "SUCCESS"
                        ),
                        "logical_page_number": (
                            logical_page_number
                        ),
                        "row_count": sheet_row_count,
                        "raw_max_row": (
                            raw_max_row
                        ),
                        "raw_max_column": (
                            raw_max_column
                        ),
                        "effective_max_row": (
                            effective_max_row
                        ),
                        "effective_max_column": (
                            effective_max_column
                        ),
                        "used_range_strategy": (
                            used_range_diagnostics[
                                "strategy"
                            ]
                        ),
                        "used_range_trimmed": (
                            used_range_diagnostics[
                                "trimmed"
                            ]
                        ),
                        "used_range_trimmed_row_count": (
                            used_range_diagnostics[
                                "trimmed_row_count"
                            ]
                        ),
                        "used_range_trimmed_column_count": (
                            used_range_diagnostics[
                                "trimmed_column_count"
                            ]
                        ),
                        "used_range_content_first_row": (
                            used_range_diagnostics[
                                "content_first_row"
                            ]
                        ),
                        "used_range_content_last_row": (
                            used_range_diagnostics[
                                "content_last_row"
                            ]
                        ),
                        "used_range_content_first_column": (
                            used_range_diagnostics[
                                "content_first_column"
                            ]
                        ),
                        "used_range_content_last_column": (
                            used_range_diagnostics[
                                "content_last_column"
                            ]
                        ),
                        "non_empty_row_count": (
                            sheet_non_empty_row_count
                        ),
                        "cell_count": sheet_cell_count,
                        "non_empty_cell_count": (
                            sheet_non_empty_cell_count
                        ),
                        "first_data_row": first_data_row,
                        "last_data_row": last_data_row,
                        "first_data_column": (
                            get_column_letter(
                                first_data_column
                            )
                            if first_data_column
                            else None
                        ),
                        "last_data_column": (
                            get_column_letter(
                                last_data_column
                            )
                            if last_data_column
                            else None
                        ),
                    }
                )

            character_count = sum(
                len(block.text)
                for block in blocks
            )

            metadata = {
                "source_format": "xlsx",
                "loader": "XLSXLoader",
                "loader_status": "SUCCESS",
                "read_only": self.read_only,
                "data_only": self.data_only,
                "keep_links": self.keep_links,
                "include_empty_cells_between_values": (
                    self.include_empty_cells_between_values
                ),
                "column_position_preserved": (
                    self.include_empty_cells_between_values
                ),
                "preserve_formulas_in_metadata": (
                    self.preserve_formulas_in_metadata
                ),
                "trim_inflated_used_range": (
                    self.trim_inflated_used_range
                ),
                "used_range_strategy": (
                    "worksheet_xml_nonempty_cell_bounds_v1"
                    if self.trim_inflated_used_range
                    else "worksheet_declared_dimension"
                ),
                "used_range_trimmed_sheet_count": (
                    used_range_trimmed_sheet_count
                ),
                "used_range_total_raw_max_row": (
                    used_range_total_raw_max_row
                ),
                "used_range_total_effective_max_row": (
                    used_range_total_effective_max_row
                ),
                "sheet_count": total_sheet_count,
                "processed_sheet_count": (
                    processed_sheet_count
                ),
                "skipped_sheet_count": (
                    skipped_sheet_count
                ),
                "empty_sheet_count": (
                    empty_sheet_count
                ),
                "row_count": total_row_count,
                "non_empty_row_count": (
                    total_non_empty_row_count
                ),
                "cell_count": total_cell_count,
                "non_empty_cell_count": (
                    total_non_empty_cell_count
                ),
                "block_count": len(blocks),
                "character_count": character_count,
                "sheets": sheet_metadata,
            }

            if self.print_summary:
                self._print_summary(
                    path=path,
                    metadata=metadata,
                )

            return Document(
                file_name=path.name,
                file_type="xlsx",
                pages=pages,
                blocks=blocks,
                chapters=[],
                sections=[],
                contents=[],
                metadata=metadata,
            )

        except XLSXLoaderError:
            raise

        except Exception as exc:
            raise XLSXLoaderError(
                f"Failed to load XLSX file "
                f"'{path.name}': {exc}"
            ) from exc

        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:
                    pass

    def _iter_rows(
        self,
        worksheet: Worksheet,
        *,
        max_row: int,
        max_column: int,
    ) -> Iterable[tuple[Any, ...]]:
        """
        顺序读取工作表行。

        重要：
            不直接依赖 worksheet.max_row / max_column。

        某些 Excel 文件只因为整列/大量空行设置过格式，
        worksheet.max_row 就可能膨胀到：

            1,048,576

        即使实际业务数据只到 202 行。

        max_row / max_column 已由 _resolve_effective_used_range()
        计算为“实际有值单元格范围”。
        """

        if (
            max_row <= 0
            or max_column <= 0
        ):

            return iter(
                ()
            )

        requested_max_row = (
            max_row
        )

        if (
            self.maximum_rows_per_sheet
            is not None
        ):

            requested_max_row = min(
                requested_max_row,
                self.maximum_rows_per_sheet,
            )

        requested_max_column = (
            max_column
        )

        if (
            self.maximum_columns_per_sheet
            is not None
        ):

            requested_max_column = min(
                requested_max_column,
                self.maximum_columns_per_sheet,
            )

        return worksheet.iter_rows(
            min_row=1,
            max_row=(
                requested_max_row
            ),
            min_col=1,
            max_col=(
                requested_max_column
            ),
        )

    # ==================================================
    # Used Range / Inflated Dimension Guard
    # ==================================================

    def _resolve_effective_used_range(
        self,
        *,
        workbook_path: Path,
        worksheet: Worksheet,
        raw_max_row: int,
        raw_max_column: int,
    ) -> tuple[
        int,
        int,
        dict[str, Any],
    ]:
        """
        返回实际需要读取的最大行/列。

        Excel 经常出现：

            raw_max_row = 1,048,551
            real_last_row = 202

        原因通常是格式/样式曾被应用到大量空白行。

        这里直接扫描对应 worksheet XML 中真正带值的 Cell：

            <c ...><v>...</v></c>
            <c ...><f>...</f></c>
            <c t="inlineStr">...</c>

        仅有样式、没有值的：

            <c r="T1048551" s="12"/>

        不计入有效 UsedRange。

        如果 XML 扫描失败，则安全回退到 worksheet 声明范围。
        """

        raw_row = max(
            int(
                raw_max_row
                or 0
            ),
            0,
        )

        raw_column = max(
            int(
                raw_max_column
                or 0
            ),
            0,
        )

        diagnostics: dict[
            str,
            Any,
        ] = {
            "strategy": (
                "worksheet_declared_dimension"
            ),
            "trimmed": (
                False
            ),
            "trimmed_row_count": (
                0
            ),
            "trimmed_column_count": (
                0
            ),
            "content_first_row": (
                None
            ),
            "content_last_row": (
                None
            ),
            "content_first_column": (
                None
            ),
            "content_last_column": (
                None
            ),
        }

        if not self.trim_inflated_used_range:

            return (
                raw_row,
                raw_column,
                diagnostics,
            )

        try:

            (
                first_row,
                last_row,
                first_column,
                last_column,
            ) = self._scan_worksheet_xml_content_bounds(
                workbook_path=(
                    workbook_path
                ),
                worksheet=(
                    worksheet
                ),
            )

        except Exception:

            diagnostics[
                "strategy"
            ] = (
                "worksheet_declared_dimension_fallback"
            )

            return (
                raw_row,
                raw_column,
                diagnostics,
            )

        diagnostics.update(
            {
                "strategy": (
                    "worksheet_xml_nonempty_cell_bounds_v1"
                ),
                "content_first_row": (
                    first_row
                ),
                "content_last_row": (
                    last_row
                ),
                "content_first_column": (
                    first_column
                ),
                "content_last_column": (
                    last_column
                ),
            }
        )

        # Truly empty worksheet.
        if (
            last_row is None
            or last_column is None
        ):

            effective_row = 0
            effective_column = 0

        else:

            effective_row = int(
                last_row
            )

            effective_column = int(
                last_column
            )

        if (
            self.maximum_rows_per_sheet
            is not None
        ):

            effective_row = min(
                effective_row,
                self.maximum_rows_per_sheet,
            )

        if (
            self.maximum_columns_per_sheet
            is not None
        ):

            effective_column = min(
                effective_column,
                self.maximum_columns_per_sheet,
            )

        row_extra = max(
            raw_row
            - effective_row,
            0,
        )

        column_extra = max(
            raw_column
            - effective_column,
            0,
        )

        row_ratio = (
            (
                raw_row
                / max(
                    effective_row,
                    1,
                )
            )
            if raw_row
            else 1.0
        )

        # Only label it as "inflated" when the gap is materially large.
        #
        # We still use XML-derived effective bounds for normal workbooks;
        # the flag is specifically diagnostic.
        trimmed = (
            row_extra
            >= self.used_range_inflation_min_extra_rows
            and row_ratio
            >= self.used_range_inflation_ratio
        )

        diagnostics[
            "trimmed"
        ] = (
            trimmed
        )

        diagnostics[
            "trimmed_row_count"
        ] = (
            row_extra
            if trimmed
            else 0
        )

        diagnostics[
            "trimmed_column_count"
        ] = (
            column_extra
            if trimmed
            else 0
        )

        return (
            effective_row,
            effective_column,
            diagnostics,
        )

    @classmethod
    def _scan_worksheet_xml_content_bounds(
        cls,
        *,
        workbook_path: Path,
        worksheet: Worksheet,
    ) -> tuple[
        int | None,
        int | None,
        int | None,
        int | None,
    ]:
        """
        流式扫描 worksheet XML，找真正含值/公式的 Cell 范围。

        不加载整个 XML 到内存。
        """

        worksheet_path = str(
            getattr(
                worksheet,
                "_worksheet_path",
                "",
            )
            or ""
        ).lstrip(
            "/"
        )

        if not worksheet_path:

            raise XLSXLoaderError(
                "Worksheet XML path is unavailable."
            )

        first_row: int | None = None
        last_row: int | None = None
        first_column: int | None = None
        last_column: int | None = None

        with zipfile.ZipFile(
            workbook_path,
            mode="r",
        ) as archive:

            with archive.open(
                worksheet_path,
                mode="r",
            ) as stream:

                for (
                    event,
                    element,
                ) in iterparse(
                    stream,
                    events=(
                        "end",
                    ),
                ):

                    if not element.tag.endswith(
                        "}c"
                    ):

                        continue

                    if not cls._xml_cell_has_content(
                        element
                    ):

                        element.clear()

                        continue

                    coordinate = str(
                        element.attrib.get(
                            "r",
                            "",
                        )
                        or ""
                    ).strip()

                    if not coordinate:

                        element.clear()

                        continue

                    try:

                        (
                            row_number,
                            column_number,
                        ) = coordinate_to_tuple(
                            coordinate
                        )

                    except Exception:

                        element.clear()

                        continue

                    first_row = (
                        row_number
                        if first_row is None
                        else min(
                            first_row,
                            row_number,
                        )
                    )

                    last_row = (
                        row_number
                        if last_row is None
                        else max(
                            last_row,
                            row_number,
                        )
                    )

                    first_column = (
                        column_number
                        if first_column is None
                        else min(
                            first_column,
                            column_number,
                        )
                    )

                    last_column = (
                        column_number
                        if last_column is None
                        else max(
                            last_column,
                            column_number,
                        )
                    )

                    element.clear()

        return (
            first_row,
            last_row,
            first_column,
            last_column,
        )

    @staticmethod
    def _xml_cell_has_content(
        cell_element: Any,
    ) -> bool:
        """
        判断 worksheet XML 的 <c> 是否包含真实内容。

        有效：
            <v>
            <f>
            inlineStr / <is><t>...</t></is>

        无效：
            只有 style / format 的空 Cell。
        """

        for child in list(
            cell_element
        ):

            tag = str(
                child.tag
            )

            if tag.endswith(
                "}f"
            ):

                return True

            if tag.endswith(
                "}v"
            ):

                return True

            if tag.endswith(
                "}is"
            ):

                for descendant in child.iter():

                    if str(
                        descendant.tag
                    ).endswith(
                        "}t"
                    ):

                        if str(
                            descendant.text
                            or ""
                        ):

                            return True

        return False

    def _extract_row(
        self,
        *,
        row: tuple[Any, ...],
        row_number: int,
    ) -> tuple[
        list[str],
        list[str],
        dict[str, str],
        int,
        int,
    ]:
        """
        提取单行。

        Returns:
            values:
                清理后的单元格文本。

            coordinates:
                与 values 对应的单元格坐标。

            formulas:
                单元格坐标到公式文本的映射。

            raw_cell_count:
                原始行单元格数量。

            non_empty_cell_count:
                非空单元格数量。
        """

        normalized_values: list[str] = []
        coordinates: list[str] = []
        formulas: dict[str, str] = {}

        raw_cell_count = len(row)
        non_empty_cell_count = 0

        for column_index, cell in enumerate(
            row,
            start=1,
        ):
            if (
                self.maximum_columns_per_sheet
                is not None
                and column_index
                > self.maximum_columns_per_sheet
            ):
                break

            value = self._get_cell_value(
                cell
            )

            normalized = self._normalize_cell_value(
                value
            )

            coordinate = self._resolve_coordinate(
                cell=cell,
                row_number=row_number,
                column_index=column_index,
            )

            if normalized:
                non_empty_cell_count += 1

            normalized_values.append(
                normalized
            )

            coordinates.append(
                coordinate
            )

            if self.preserve_formulas_in_metadata:
                formula = self._extract_formula(
                    cell
                )

                if formula is not None:
                    formulas[
                        coordinate
                    ] = formula

        (
            normalized_values,
            coordinates,
        ) = self._trim_trailing_empty_cells(
            values=normalized_values,
            coordinates=coordinates,
        )

        if not self.include_empty_cells_between_values:
            filtered_values: list[str] = []
            filtered_coordinates: list[str] = []

            for value, coordinate in zip(
                normalized_values,
                coordinates,
            ):
                if not value:
                    continue

                filtered_values.append(value)
                filtered_coordinates.append(
                    coordinate
                )

            normalized_values = filtered_values
            coordinates = filtered_coordinates

        if not any(normalized_values):
            return (
                [],
                [],
                {},
                raw_cell_count,
                non_empty_cell_count,
            )

        valid_formula_coordinates = set(
            coordinates
        )

        formulas = {
            coordinate: formula
            for coordinate, formula
            in formulas.items()
            if coordinate
            in valid_formula_coordinates
        }

        return (
            normalized_values,
            coordinates,
            formulas,
            raw_cell_count,
            non_empty_cell_count,
        )

    @staticmethod
    def _get_cell_value(
        cell: Any,
    ) -> Any:

        return getattr(
            cell,
            "value",
            None,
        )

    @staticmethod
    def _extract_formula(
        cell: Any,
    ) -> str | None:

        if not isinstance(
            cell,
            Cell,
        ):
            return None

        if cell.data_type != "f":
            return None

        value = cell.value

        if value is None:
            return None

        return str(value)

    @classmethod
    def _normalize_cell_value(
        cls,
        value: Any,
    ) -> str:

        if value is None:
            return ""

        if isinstance(value, bool):
            return (
                "TRUE"
                if value
                else "FALSE"
            )

        if isinstance(value, datetime):
            return value.isoformat(
                sep=" ",
                timespec="seconds",
            )

        if isinstance(value, date):
            return value.isoformat()

        if isinstance(value, time):
            return value.isoformat(
                timespec="seconds"
            )

        if isinstance(value, timedelta):
            return str(value)

        if isinstance(value, Decimal):
            return cls._normalize_decimal(
                value
            )

        if isinstance(value, int):
            return str(value)

        if isinstance(value, float):
            return cls._normalize_float(
                value
            )

        if isinstance(value, bytes):
            return value.decode(
                "utf-8",
                errors="replace",
            )

        return cls._normalize_text(
            str(value)
        )

    @staticmethod
    def _normalize_decimal(
        value: Decimal,
    ) -> str:

        normalized = value.normalize()

        return format(
            normalized,
            "f",
        )

    @staticmethod
    def _normalize_float(
        value: float,
    ) -> str:

        if math.isnan(value):
            return "NaN"

        if math.isinf(value):
            return (
                "Infinity"
                if value > 0
                else "-Infinity"
            )

        if value.is_integer():
            return str(
                int(value)
            )

        return format(
            value,
            ".15g",
        )

    @staticmethod
    def _normalize_text(
        value: str,
    ) -> str:

        normalized = value.replace(
            "\u3000",
            " ",
        )

        normalized = normalized.replace(
            "\xa0",
            " ",
        )

        for character in (
            "\u200b",
            "\u200c",
            "\u200d",
            "\u2060",
            "\ufeff",
        ):
            normalized = normalized.replace(
                character,
                "",
            )

        normalized = normalized.replace(
            "\r\n",
            "\n",
        )

        normalized = normalized.replace(
            "\r",
            "\n",
        )

        lines = [
            " ".join(line.split())
            for line in normalized.splitlines()
            if line.strip()
        ]

        return " / ".join(
            lines
        ).strip()

    @staticmethod
    def _trim_trailing_empty_cells(
        *,
        values: list[str],
        coordinates: list[str],
    ) -> tuple[list[str], list[str]]:
        """
        仅删除最右侧连续空单元格。

        重要：
            行首空 Cell 必须保留，否则会破坏真实列位置。

        Example:

            ["", "", "A", "", "B", "", ""]

        ->

            ["", "", "A", "", "B"]

        这样第一个有效值仍然明确位于 C 列。
        """

        if not values:
            return [], []

        end = len(values)

        while (
            end > 0
            and not values[end - 1]
        ):
            end -= 1

        if end == 0:
            return [], []

        return (
            values[:end],
            coordinates[:end],
        )

    @staticmethod
    def _trim_row_boundaries(
        *,
        values: list[str],
        coordinates: list[str],
    ) -> tuple[list[str], list[str]]:
        """
        兼容旧私有方法名。

        当前语义已经改为：
            仅裁剪行尾空 Cell。

        不再删除行首空 Cell。
        """

        return XLSXLoader._trim_trailing_empty_cells(
            values=values,
            coordinates=coordinates,
        )

    @staticmethod
    def _resolve_coordinate(
        *,
        cell: Any,
        row_number: int,
        column_index: int,
    ) -> str:

        coordinate = getattr(
            cell,
            "coordinate",
            None,
        )

        if coordinate:
            return str(
                coordinate
            )

        return (
            f"{get_column_letter(column_index)}"
            f"{row_number}"
        )

    def _should_include_sheet(
        self,
        sheet_state: str,
    ) -> bool:

        if sheet_state == "visible":
            return True

        if sheet_state == "hidden":
            return self.include_hidden_sheets

        if sheet_state == "veryHidden":
            return (
                self.include_very_hidden_sheets
            )

        return False

    @staticmethod
    def _get_sheet_state(
        worksheet: Worksheet,
    ) -> str:

        state = getattr(
            worksheet,
            "sheet_state",
            "visible",
        )

        return str(
            state
        )

    @classmethod
    def _validate_path(
        cls,
        file_path: str | Path,
    ) -> Path:

        if file_path is None:
            raise ValueError(
                "file_path cannot be None."
            )

        if not str(
            file_path
        ).strip():
            raise ValueError(
                "file_path cannot be empty."
            )

        path = Path(
            file_path
        ).expanduser()

        if not path.exists():
            raise FileNotFoundError(
                f"XLSX file not found: {path}"
            )

        if not path.is_file():
            raise IsADirectoryError(
                f"Input path is not a file: {path}"
            )

        if path.name.startswith("~$"):
            raise ValueError(
                f"Temporary Excel file is not supported: "
                f"{path.name}"
            )

        if path.suffix.lower() not in (
            cls.SUPPORTED_EXTENSIONS
        ):
            supported = ", ".join(
                sorted(
                    cls.SUPPORTED_EXTENSIONS
                )
            )

            raise ValueError(
                f"XLSXLoader only accepts: {supported}. "
                f"Received: "
                f"{path.suffix or '<no extension>'}"
            )

        return path

    @staticmethod
    def _print_summary(
        *,
        path: Path,
        metadata: dict[str, Any],
    ) -> None:

        print()
        print("===== XLSX Loader =====")
        print(
            f"File            : {path.name}"
        )
        print(
            "Sheets          :",
            metadata["sheet_count"],
        )
        print(
            "Processed sheets:",
            metadata[
                "processed_sheet_count"
            ],
        )
        print(
            "Skipped sheets  :",
            metadata[
                "skipped_sheet_count"
            ],
        )
        print(
            "Empty sheets    :",
            metadata[
                "empty_sheet_count"
            ],
        )
        print(
            "Rows            :",
            metadata["row_count"],
        )
        print(
            "Non-empty rows  :",
            metadata[
                "non_empty_row_count"
            ],
        )
        print(
            "Blocks          :",
            metadata["block_count"],
        )
        print(
            "Characters      :",
            metadata["character_count"],
        )
        print("=======================")
        print()

# ======================================================
# Compatibility Alias
# ======================================================
#
# 兼容：
#
#     from app.loader.excel_loader import ExcelLoader
#
# 与：
#
#     from app.loader.xlsx_loader import XLSXLoader
#
# 使用同一实现。

ExcelLoader = XLSXLoader